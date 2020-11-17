# sendDuesReminders.py

import smtplib
import openpyxl
import sys

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_row
latestMonth = sheet.cell(row=1, column=lastCol).value

unpaidMembers = {}
for r in range(2, sheet.max_row):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

my_email = sys.argv[1]
my_pass = sys.argv[2]

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(my_email, my_pass)

for name, email in unpaidMembers.items():
    body = f'Subject: Overdue payment from {latestMonth}. \nHello, {name}\nAccording to our records, you have not ' \
           f'paid off your fee for {latestMonth}. The amount needs to be paid as soon as possible\nHowever, However, ' \
           f'if, in the meantime, you have already settled the amount, please ignore this reminder.'
    print(f'Sending an e-mail to {email}')
    '''sendmailStatus = smtpObj.sendmail(my_email, email, body)

    if sendmailStatus != {}:
        print(f'An ERROR occurred while sending mail to {email} {sendmailStatus}')'''
smtpObj.quit()
